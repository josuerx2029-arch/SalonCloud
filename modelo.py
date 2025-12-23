import sqlite3
import pandas as pd
from datetime import datetime, timedelta
import os
import hashlib  # Para encriptar contraseñas
import shutil   # Para copiar archivos (Backup y Logo)
import json     # Para guardar la configuración de campos obligatorios

# Importaciones opcionales para Excel y PDF
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

class SistemaSalonDB:
    def __init__(self, db_name="salon_sistema_pro.db"):
        self.db_name = db_name
        self.inicializar_tablas()
        self.migrar_db_a_iso() # Ejecuta la corrección de fechas automáticamente al iniciar

    def conectar(self):
        return sqlite3.connect(self.db_name)

    # ==========================================
    #  HELPERS: TRADUCTORES DE FECHAS
    # ==========================================
    def f_to_iso(self, fecha_ui):
        """Convierte '21-12-25' (App) a '2025-12-21' (Base de Datos)"""
        try:
            return datetime.strptime(fecha_ui, "%d-%m-%y").strftime("%Y-%m-%d")
        except:
            return fecha_ui # Si falla, retorna original

    def f_to_ui(self, fecha_iso):
        """Convierte '2025-12-21' (Base de Datos) a '21-12-25' (App)"""
        try:
            return datetime.strptime(fecha_iso, "%Y-%m-%d").strftime("%d-%m-%y")
        except:
            return fecha_iso

    # ==========================================
    #  INICIALIZACIÓN Y MIGRACIÓN
    # ==========================================
    def inicializar_tablas(self):
        with self.conectar() as conn:
            c = conn.cursor()
            
            # --- NUEVA TABLA UNIFICADA DE TERCEROS ---
            # Reemplaza a clientes, proveedores y profesionales individuales.
            # Permite tener roles mixtos (Ej: Empleado que también es Cliente).
            c.execute('''CREATE TABLE IF NOT EXISTS terceros (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                
                -- Identificación y Nombres
                doc_id TEXT UNIQUE,       -- Cédula, NIT o RUT
                nombre1 TEXT,             -- Primer Nombre
                nombre2 TEXT,             -- Segundo Nombre
                apellido1 TEXT,           -- Primer Apellido
                apellido2 TEXT,           -- Segundo Apellido
                nombre_completo TEXT,     -- Campo calculado (Concatenación) para búsquedas rápidas
                
                -- Datos de Contacto
                direccion TEXT,
                telefono TEXT,
                email TEXT,
                ciudad TEXT,
                fecha_nacimiento TEXT,
                
                -- ROLES (Banderas booleanas 0/1)
                es_cliente INTEGER DEFAULT 0,
                es_proveedor INTEGER DEFAULT 0,
                es_empleado INTEGER DEFAULT 0,
                
                -- Datos Específicos de Empleado
                comision REAL DEFAULT 0,
                color_agenda TEXT,
                servicios_asignados TEXT, -- "Corte,Tinte" o "TODOS"
                
                -- Metadatos
                fecha_registro TEXT,
                notas_internas TEXT
            )''')

            # --- MIGRACIÓN DE DATOS ANTIGUOS (SI EXISTEN) ---
            # Si existían tablas viejas, idealmente aquí se haría un script de migración.
            # Para este código, asumimos que si no existen las tablas viejas, usamos la nueva lógica.
            # Mantenemos las definiciones viejas COMENTADAS o como respaldo si se requiere compatibilidad,
            # pero el sistema ahora priorizará 'terceros'.
            
            # 1. Tablas Maestras de Configuración y Servicios
            c.execute('''CREATE TABLE IF NOT EXISTS servicios (
                id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT UNIQUE, duracion_min INTEGER, precio REAL)''')
            c.execute('''CREATE TABLE IF NOT EXISTS medios_pago (nombre TEXT UNIQUE)''')
            c.execute('''CREATE TABLE IF NOT EXISTS configuracion (clave TEXT PRIMARY KEY, valor TEXT)''')
            
            # TABLA USUARIOS (LOGIN)
            c.execute('''CREATE TABLE IF NOT EXISTS usuarios (
                usuario TEXT PRIMARY KEY, 
                password_hash TEXT, 
                rol TEXT DEFAULT 'admin')''')

            # --- TABLAS INVENTARIO ---
            c.execute('''CREATE TABLE IF NOT EXISTS productos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, 
                nombre TEXT UNIQUE, 
                precio REAL, 
                stock INTEGER,
                codigo_barras TEXT)''')
            
            c.execute('''CREATE TABLE IF NOT EXISTS ventas_productos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, 
                cita_id INTEGER, 
                producto_id INTEGER, 
                cantidad INTEGER, 
                precio_unitario REAL,
                fecha TEXT,
                FOREIGN KEY(cita_id) REFERENCES citas(id))''')

            # --- TABLAS DE COMPRAS (Integradas con Terceros) ---
            c.execute('''CREATE TABLE IF NOT EXISTS compras (
                id INTEGER PRIMARY KEY AUTOINCREMENT, 
                proveedor_id INTEGER, -- Referencia a tabla terceros
                fecha TEXT, 
                total REAL, 
                metodo_pago TEXT, 
                estado TEXT, -- 'Pagado' o 'Pendiente'
                observacion TEXT)''')

            c.execute('''CREATE TABLE IF NOT EXISTS detalle_compras (
                id INTEGER PRIMARY KEY AUTOINCREMENT, 
                compra_id INTEGER, 
                producto_id INTEGER, 
                cantidad INTEGER, 
                costo_unitario REAL,
                FOREIGN KEY(compra_id) REFERENCES compras(id))''')
                
            c.execute('''CREATE TABLE IF NOT EXISTS abonos_proveedores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                compra_id INTEGER,
                monto REAL,
                fecha TEXT,
                metodo TEXT)''')

            # 2. Operativas (Citas referencia a Terceros)
            c.execute('''CREATE TABLE IF NOT EXISTS citas (
                id INTEGER PRIMARY KEY AUTOINCREMENT, 
                cliente_id INTEGER,      -- Referencia a terceros (es_cliente=1)
                profesional_id INTEGER,  -- Referencia a terceros (es_empleado=1)
                servicio_id INTEGER,
                fecha TEXT, hora_inicio TEXT, hora_fin TEXT, estado TEXT DEFAULT 'Pendiente', 
                precio_final REAL, descuento REAL DEFAULT 0, nomina_pagada INTEGER DEFAULT 0)''')
            
            # 3. Financieras
            c.execute('''CREATE TABLE IF NOT EXISTS pagos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, cita_id INTEGER, metodo TEXT, monto REAL, fecha TEXT, hora TEXT, descripcion_extra TEXT)''')
            c.execute('''CREATE TABLE IF NOT EXISTS gastos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, fecha TEXT, tipo TEXT, categoria TEXT, descripcion TEXT, metodo TEXT, valor REAL)''')
            
            # Prestamos (Referencia a Terceros Empleados)
            c.execute('''CREATE TABLE IF NOT EXISTS prestamos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, profesional_id INTEGER, monto REAL, fecha TEXT, estado TEXT DEFAULT 'Pendiente', descripcion TEXT)''')
            c.execute('''CREATE TABLE IF NOT EXISTS abonos_prestamos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, prestamo_id INTEGER, valor REAL, fecha TEXT, descripcion TEXT, metodo TEXT)''')
            
            # 4. Control
            c.execute('''CREATE TABLE IF NOT EXISTS bloqueos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, profesional_id INTEGER, fecha TEXT, hora_inicio TEXT, hora_fin TEXT, motivo TEXT)''')
            
            # 5. AUDITORÍA (Logs de seguridad)
            c.execute('''CREATE TABLE IF NOT EXISTS auditoria (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT,
                hora TEXT,
                accion TEXT,
                detalle TEXT,
                usuario TEXT DEFAULT 'Admin'
            )''')

            # Semilla de Datos (Solo si la base está vacía)
            # Creamos empleados por defecto en la tabla TERCEROS si no existen
            if c.execute("SELECT count(*) FROM terceros WHERE es_empleado=1").fetchone()[0] == 0:
                c.execute('''INSERT INTO terceros (nombre_completo, nombre1, comision, servicios_asignados, es_empleado, fecha_registro) 
                             VALUES ('Andrea', 'Andrea', 50, 'TODOS', 1, ?)''', (datetime.now().strftime("%Y-%m-%d"),))
                c.execute('''INSERT INTO terceros (nombre_completo, nombre1, comision, servicios_asignados, es_empleado, fecha_registro) 
                             VALUES ('Lennys', 'Lennys', 50, 'TODOS', 1, ?)''', (datetime.now().strftime("%Y-%m-%d"),))

            if c.execute("SELECT count(*) FROM servicios").fetchone()[0] == 0:
                c.executemany("INSERT INTO servicios (nombre, duracion_min, precio) VALUES (?,?,?)", 
                             [('Cejas 3D', 60, 50000), ('Pestañas', 60, 80000), ('Diseño', 30, 30000)])
            if c.execute("SELECT count(*) FROM medios_pago").fetchone()[0] == 0:
                c.executemany("INSERT INTO medios_pago (nombre) VALUES (?)", 
                             [('Efectivo',), ('Nequi',), ('Daviplata',), ('Tarjeta',), ('CREDITO',)])
            c.execute("INSERT OR IGNORE INTO configuracion (clave, valor) VALUES (?,?)", 
                      ('msg_wa', '{saludo} {cliente}, paso a confirmar su cita para {dia_relativo} a las {hora} con {profesional} para {servicio}. Valor: {precio}.'))
            
            # Configuración por defecto para Campos Obligatorios de Terceros
            # doc_id, n1, a1 son obligatorios por lógica interna, pero el usuario puede definir otros
            try:
                config_terceros = json.dumps({"doc_id": True, "telefono": False, "email": False, "direccion": False})
                c.execute("INSERT OR IGNORE INTO configuracion (clave, valor) VALUES (?,?)", ('campos_obligatorios', config_terceros))
            except: pass

            # CREAR ADMIN POR DEFECTO SI NO EXISTE
            if c.execute("SELECT count(*) FROM usuarios").fetchone()[0] == 0:
                # Contraseña por defecto: 1234
                pass_defecto = hashlib.sha256("1234".encode()).hexdigest()
                c.execute("INSERT INTO usuarios (usuario, password_hash) VALUES (?,?)", ("admin", pass_defecto))

            conn.commit()

    def migrar_db_a_iso(self):
        """Revisa si hay fechas viejas (dd-mm-yy) y las pasa a (yyyy-mm-dd)"""
        tablas_fechas = {
            'citas': 'fecha', 
            'pagos': 'fecha', 
            'gastos': 'fecha', 
            'prestamos': 'fecha', 
            'abonos_prestamos': 'fecha', 
            'bloqueos': 'fecha',
            'terceros': 'fecha_registro',
            'ventas_productos': 'fecha',
            'compras': 'fecha'
        }
        with self.conectar() as conn:
            c = conn.cursor()
            for tabla, col in tablas_fechas.items():
                try:
                    sql_check = f"SELECT count(*) FROM {tabla} WHERE {col} LIKE '__-__-__'"
                    if c.execute(sql_check).fetchone()[0] > 0:
                        c.execute(f"UPDATE {tabla} SET {col} = '20'||substr({col},7,2)||'-'||substr({col},4,2)||'-'||substr({col},1,2) WHERE {col} LIKE '__-__-__' ")
                except: pass
            conn.commit()

    # ==========================================
    #  NUEVO: GESTIÓN UNIFICADA DE TERCEROS
    # ==========================================
    
    def guardar_tercero(self, datos):
        """
        Recibe un diccionario con los datos del tercero y lo guarda o actualiza.
        Calcula el nombre completo automáticamente.
        """
        try:
            # Extraer datos básicos
            n1 = datos.get('n1', '').strip()
            n2 = datos.get('n2', '').strip()
            a1 = datos.get('a1', '').strip()
            a2 = datos.get('a2', '').strip()
            
            # Construir Nombre Completo (Lógica de negocio)
            # Ej: "Franklin Jose Perez Lopez"
            partes = [n1, n2, a1, a2]
            nombre_completo = " ".join([p for p in partes if p])
            
            # Roles (Convertir bool a int 0/1)
            es_cli = 1 if datos.get('es_cliente') else 0
            es_prov = 1 if datos.get('es_proveedor') else 0
            es_emp = 1 if datos.get('es_empleado') else 0
            
            valores = (
                datos.get('doc_id'),
                n1, n2, a1, a2,
                nombre_completo,
                datos.get('direccion'),
                datos.get('telefono'),
                datos.get('email'),
                datos.get('ciudad'),
                es_cli, es_prov, es_emp,
                datos.get('comision', 0),
                datos.get('color', ''),
                datos.get('servicios', ''),
                datetime.now().strftime("%Y-%m-%d")
            )
            
            id_existente = datos.get('id')
            
            with self.conectar() as conn:
                if id_existente:
                    # UPDATE
                    sql = '''UPDATE terceros SET 
                             doc_id=?, nombre1=?, nombre2=?, apellido1=?, apellido2=?, nombre_completo=?,
                             direccion=?, telefono=?, email=?, ciudad=?, 
                             es_cliente=?, es_proveedor=?, es_empleado=?, 
                             comision=?, color_agenda=?, servicios_asignados=?
                             WHERE id=?'''
                    # Nota: No actualizamos fecha_registro en update
                    params = valores[:-1] + (id_existente,)
                    conn.execute(sql, params)
                    accion = f"Actualizado Tercero: {nombre_completo}"
                else:
                    # INSERT
                    sql = '''INSERT INTO terceros (
                             doc_id, nombre1, nombre2, apellido1, apellido2, nombre_completo,
                             direccion, telefono, email, ciudad, 
                             es_cliente, es_proveedor, es_empleado, 
                             comision, color_agenda, servicios_asignados, fecha_registro)
                             VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''
                    conn.execute(sql, valores)
                    accion = f"Creado Tercero: {nombre_completo}"
            
            self.registrar_auditoria("TERCEROS", accion)
            return True, "Datos guardados correctamente"
            
        except sqlite3.IntegrityError:
            return False, "Error: El Documento ID ya existe en la base de datos."
        except Exception as e:
            return False, f"Error al guardar: {str(e)}"

    def buscar_tercero_general(self, texto_busqueda):
        """Busca en la tabla terceros por nombre, documento o teléfono"""
        filtro = f"%{texto_busqueda}%"
        sql = '''SELECT * FROM terceros 
                 WHERE nombre_completo LIKE ? OR doc_id LIKE ? OR telefono LIKE ?
                 ORDER BY nombre_completo LIMIT 50'''
        with self.conectar() as conn:
            conn.row_factory = sqlite3.Row # Para acceder por nombre de columna
            return conn.execute(sql, (filtro, filtro, filtro)).fetchall()

    def traer_tercero_por_id(self, id_tercero):
        with self.conectar() as conn:
            conn.row_factory = sqlite3.Row
            return conn.execute("SELECT * FROM terceros WHERE id=?", (id_tercero,)).fetchone()

    # --- MÉTODOS DE COMPATIBILIDAD (Adaptados a la nueva estructura) ---

    def buscar_cliente(self, texto):
        """Usado en Agendar y Recepción. Busca solo si es_cliente=1"""
        filtro = f"%{texto}%"
        with self.conectar() as conn:
            # Retorna ID, Nombre, Telefono
            res = conn.execute('''SELECT id, nombre_completo, telefono FROM terceros 
                                  WHERE es_cliente=1 AND (nombre_completo LIKE ? OR telefono LIKE ? OR doc_id LIKE ?)''', 
                                  (filtro, filtro, filtro)).fetchone()
            return res

    def get_listas(self):
        """Retorna lista de Nombres de Empleados y Servicios"""
        with self.conectar() as conn:
            # Profesionales: es_empleado = 1
            pros = [x[0] for x in conn.execute("SELECT nombre_completo FROM terceros WHERE es_empleado=1").fetchall()]
            servs = [x[0] for x in conn.execute("SELECT nombre FROM servicios").fetchall()]
        return pros, servs

    def traer_proveedores(self):
        """Retorna lista de proveedores para el módulo de compras"""
        with self.conectar() as conn:
            return conn.execute("SELECT id, nombre_completo, telefono, direccion FROM terceros WHERE es_proveedor=1 ORDER BY nombre_completo").fetchall()

    def traer_tabla_completa(self, tipo):
        """Retorna datos para las tablas de configuración"""
        with self.conectar() as conn:
            if tipo == 'servicios':
                return conn.execute("SELECT nombre, duracion_min, precio FROM servicios").fetchall()
            elif tipo == 'profesionales':
                # Ahora leemos de terceros con flag empleado
                return conn.execute("SELECT nombre_completo, comision, servicios_asignados FROM terceros WHERE es_empleado=1").fetchall()
            elif tipo == 'clientes':
                # Ahora leemos de terceros con flag cliente
                rows = conn.execute("SELECT id, nombre_completo, telefono, fecha_registro FROM terceros WHERE es_cliente=1 ORDER BY id DESC LIMIT 100").fetchall()
                return [(r[0], r[1], r[2], self.f_to_ui(r[3])) for r in rows]
        return []

    # ==========================================
    #  SISTEMA DE CONFIGURACIÓN (CAMPOS OBLIGATORIOS)
    # ==========================================
    def guardar_config_campos(self, dict_campos):
        """Guarda qué campos son obligatorios para terceros"""
        try:
            json_val = json.dumps(dict_campos)
            with self.conectar() as conn:
                conn.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES (?,?)", ('campos_obligatorios', json_val))
            return True, "Configuración actualizada"
        except Exception as e: return False, str(e)

    def traer_config_campos(self):
        """Retorna diccionario con booleanos de campos obligatorios"""
        try:
            with self.conectar() as conn:
                res = conn.execute("SELECT valor FROM configuracion WHERE clave='campos_obligatorios'").fetchone()
                if res: return json.loads(res[0])
        except: pass
        # Valor por defecto
        return {"doc_id": True, "telefono": False, "email": False, "direccion": False}

    # ==========================================
    #  SISTEMA DE AUDITORÍA
    # ==========================================
    def registrar_auditoria(self, accion, detalle):
        try:
            with self.conectar() as conn:
                conn.execute("INSERT INTO auditoria (fecha, hora, accion, detalle) VALUES (?,?,?,?)", 
                             (datetime.now().strftime("%Y-%m-%d"), datetime.now().strftime("%H:%M:%S"), accion, detalle))
        except Exception as e:
            print(f"Error Audit: {e}")

    def traer_auditoria(self):
        with self.conectar() as conn:
            rows = conn.execute("SELECT fecha, hora, accion, detalle, usuario FROM auditoria ORDER BY id DESC LIMIT 100").fetchall()
            return [(self.f_to_ui(r[0]), r[1], r[2], r[3], r[4]) for r in rows]

    # ==========================================
    #  SEGURIDAD Y BACKUP
    # ==========================================
    def validar_login(self, usuario, password_texto):
        pass_hash = hashlib.sha256(password_texto.encode()).hexdigest()
        with self.conectar() as conn:
            data = conn.execute("SELECT usuario FROM usuarios WHERE usuario=? AND password_hash=?", (usuario, pass_hash)).fetchone()
            if data:
                self.registrar_auditoria("LOGIN", f"Ingreso exitoso: {usuario}")
                return True
            else:
                self.registrar_auditoria("LOGIN_FAIL", f"Intento fallido: {usuario}")
                return False

    def generar_backup_db(self, carpeta_destino):
        try:
            if not os.path.exists(self.db_name):
                return False, "No existe la base de datos aún."
            timestamp = datetime.now().strftime("%Y_%m_%d_%H%M%S")
            nombre_backup = f"RESPALDO_SALON_{timestamp}.db"
            ruta_completa = os.path.join(carpeta_destino, nombre_backup)
            shutil.copy2(self.db_name, ruta_completa)
            self.registrar_auditoria("BACKUP", f"Copia creada: {nombre_backup}")
            return True, f"Respaldo creado correctamente en:\n{ruta_completa}"
        except Exception as e:
            return False, f"Error al crear respaldo: {str(e)}"

    def cambiar_clave_usuario(self, usuario, nueva_clave):
        try:
            nuevo_hash = hashlib.sha256(nueva_clave.encode()).hexdigest()
            with self.conectar() as conn:
                conn.execute("UPDATE usuarios SET password_hash=? WHERE usuario=?", (nuevo_hash, usuario))
            self.registrar_auditoria("CAMBIO_CLAVE", f"Usuario {usuario} cambió su clave")
            return True, "Clave actualizada"
        except Exception as e: return False, str(e)

    # ==========================================
    #  GESTIÓN PERFIL EMPRESA
    # ==========================================
    def guardar_datos_empresa(self, nombre, nit, direccion, telefono, ruta_logo_origen=None):
        try:
            with self.conectar() as conn:
                # Guardamos los datos de texto en la tabla configuracion
                conn.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('emp_nombre', ?)", (nombre,))
                conn.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('emp_nit', ?)", (nit,))
                conn.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('emp_dir', ?)", (direccion,))
                conn.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('emp_tel', ?)", (telefono,))
            
            # Gestionar el logo (copiar archivo)
            if ruta_logo_origen and os.path.exists(ruta_logo_origen):
                shutil.copy2(ruta_logo_origen, "logo.png")
                
            self.registrar_auditoria("CONFIG_EMPRESA", "Se actualizaron datos de la empresa")
            return True, "Datos y Logo guardados correctamente"
        except Exception as e:
            return False, f"Error al guardar: {str(e)}"

    def traer_datos_empresa(self):
        """Retorna un diccionario con los datos, o vacíos si no existen"""
        datos = {'nombre': 'MI SALÓN DE BELLEZA', 'nit': '', 'dir': '', 'tel': ''}
        try:
            with self.conectar() as conn:
                rows = conn.execute("SELECT clave, valor FROM configuracion WHERE clave LIKE 'emp_%'").fetchall()
                for k, v in rows:
                    if k == 'emp_nombre': datos['nombre'] = v
                    elif k == 'emp_nit': datos['nit'] = v
                    elif k == 'emp_dir': datos['dir'] = v
                    elif k == 'emp_tel': datos['tel'] = v
        except: pass
        return datos

    # ==========================================
    #  MÓDULO INVENTARIO Y SERVICIOS
    # ==========================================
    def crear_item_unificado(self, tipo, nombre, costo_precio, extra_data):
        try:
            with self.conectar() as conn:
                if tipo == 'SERVICIO':
                    conn.execute("INSERT INTO servicios (nombre, duracion_min, precio) VALUES (?,?,?)",
                                 (nombre, int(extra_data), float(costo_precio)))
                else:
                    conn.execute("INSERT INTO productos (nombre, precio, stock) VALUES (?,?,?)",
                                 (nombre, float(costo_precio), int(extra_data)))
            return True, f"{tipo} creado con éxito"
        except Exception as e: return False, str(e)

    def crear_producto(self, nombre, precio, stock):
        return self.crear_item_unificado('PRODUCTO', nombre, precio, stock)

    def eliminar_producto(self, id_prod):
        try:
            with self.conectar() as conn: conn.execute("DELETE FROM productos WHERE id=?", (id_prod,))
            return True, "Eliminado"
        except Exception as e: return False, str(e)

    def traer_productos(self):
        with self.conectar() as conn:
            return conn.execute("SELECT id, nombre, precio, stock FROM productos ORDER BY nombre").fetchall()

    def descontar_stock(self, id_prod, cantidad):
        try:
            with self.conectar() as conn:
                curr = conn.execute("SELECT stock FROM productos WHERE id=?", (id_prod,)).fetchone()
                if not curr: return False
                nuevo_stock = curr[0] - cantidad
                conn.execute("UPDATE productos SET stock=? WHERE id=?", (nuevo_stock, id_prod))
            return True
        except: return False

    # ==========================================
    #  MÓDULO COMPRAS (ADAPTADO A TERCEROS)
    # ==========================================
    def crear_proveedor(self, nombre, tel, direccion):
        # Mantenido para compatibilidad, pero redirige a guardar_tercero
        data = {
            'n1': nombre, # Asumimos nombre simple
            'telefono': tel,
            'direccion': direccion,
            'es_proveedor': True
        }
        return self.guardar_tercero(data)

    def eliminar_proveedor(self, id_prov):
        # En realidad elimina el tercero o le quita el flag. Aquí borramos por simplicidad
        try:
            with self.conectar() as conn: conn.execute("DELETE FROM terceros WHERE id=?", (id_prov,))
            return True, "Eliminado"
        except Exception as e: return False, str(e)

    def registrar_compra(self, id_prov, items, metodo_pago, total, observacion="Compra Inventario"):
        hoy_iso = datetime.now().strftime("%Y-%m-%d")
        estado = "Pendiente" if metodo_pago == "CREDITO" else "Pagado"
        conn = self.conectar()
        try:
            c = conn.cursor()
            c.execute("INSERT INTO compras (proveedor_id, fecha, total, metodo_pago, estado, observacion) VALUES (?,?,?,?,?,?)",
                      (id_prov, hoy_iso, total, metodo_pago, estado, observacion))
            id_compra = c.lastrowid

            for prod_id, cant, costo in items:
                c.execute("INSERT INTO detalle_compras (compra_id, producto_id, cantidad, costo_unitario) VALUES (?,?,?,?)",
                          (id_compra, prod_id, cant, costo))
                curr_stock = c.execute("SELECT stock FROM productos WHERE id=?", (prod_id,)).fetchone()
                if curr_stock:
                    nuevo_stock = curr_stock[0] + int(cant)
                    c.execute("UPDATE productos SET stock=? WHERE id=?", (nuevo_stock, prod_id))

            if metodo_pago != "CREDITO":
                desc_gasto = f"Compra ID:{id_compra} Prov:{id_prov}"
                c.execute("INSERT INTO gastos (fecha, tipo, categoria, descripcion, metodo, valor) VALUES (?,?,?,?,?,?)",
                          (hoy_iso, "COSTO", "Compra Mercancia", desc_gasto, metodo_pago, total))
            
            conn.commit()
            return True, "Compra registrada y Stock actualizado"
        except Exception as e:
            conn.rollback()
            return False, str(e)
        finally:
            conn.close()

    def traer_cuentas_por_pagar_proveedores(self):
        sql = '''
        SELECT c.id, t.nombre_completo, c.fecha, c.total, 
               (SELECT IFNULL(SUM(monto),0) FROM abonos_proveedores WHERE compra_id = c.id) as pagado
        FROM compras c 
        JOIN terceros t ON c.proveedor_id = t.id 
        WHERE c.metodo_pago = 'CREDITO' AND c.estado = 'Pendiente'
        '''
        data = []
        with self.conectar() as conn:
            rows = conn.execute(sql).fetchall()
            for r in rows:
                id_c, nom, fec, tot, pagado = r
                saldo = tot - pagado
                if saldo > 0:
                    data.append((id_c, self.f_to_ui(fec), nom, f"${tot:,.0f}", f"${saldo:,.0f}"))
        return data

    def abonar_proveedor(self, id_compra, monto, metodo):
        hoy_iso = datetime.now().strftime("%Y-%m-%d")
        conn = self.conectar()
        try:
            c = conn.cursor()
            c.execute("INSERT INTO abonos_proveedores (compra_id, monto, fecha, metodo) VALUES (?,?,?,?)",
                      (id_compra, monto, hoy_iso, metodo))
            c.execute("INSERT INTO gastos (fecha, tipo, categoria, descripcion, metodo, valor) VALUES (?,?,?,?,?,?)",
                          (hoy_iso, "GASTO", "Pago Proveedor", f"Abono Compra #{id_compra}", metodo, monto))
            row = c.execute("SELECT total FROM compras WHERE id=?", (id_compra,)).fetchone()
            total_orig = row[0]
            pagado = c.execute("SELECT SUM(monto) FROM abonos_proveedores WHERE compra_id=?", (id_compra,)).fetchone()[0]
            if pagado >= total_orig:
                c.execute("UPDATE compras SET estado='Pagado' WHERE id=?", (id_compra,))
            conn.commit()
            return True, "Abono registrado correctamente"
        except Exception as e:
            conn.rollback()
            return False, str(e)
        finally:
            conn.close()

    def traer_reporte_compras(self, f1, f2):
        d1 = self.f_to_iso(f1); d2 = self.f_to_iso(f2)
        # JOIN con terceros
        sql = '''SELECT c.id, c.fecha, t.nombre_completo, c.total, c.metodo_pago, c.estado 
                 FROM compras c JOIN terceros t ON c.proveedor_id = t.id
                 WHERE c.fecha BETWEEN ? AND ? ORDER BY c.fecha DESC'''
        with self.conectar() as conn:
            rows = conn.execute(sql, (d1, d2)).fetchall()
            return [(r[0], self.f_to_ui(r[1]), r[2], r[3], r[4], r[5]) for r in rows]

    # ==========================================
    #  MÓDULO FINANCIERO (ADAPTADO A TERCEROS)
    # ==========================================
    def procesar_cobro(self, ids_str, pagos, descuento_dinero, cliente_nombre, productos_carrito=None):
        if productos_carrito is None: productos_carrito = []
        ids = ids_str.split(',')
        id_principal = ids[0]
        conn = self.conectar()
        hoy_iso = datetime.now().strftime("%Y-%m-%d")
        try:
            c = conn.cursor()
            for i in ids: 
                c.execute("UPDATE citas SET estado='Pagado', precio_final = (SELECT precio_final FROM citas WHERE id=?) WHERE id=?", (i, i))
            
            for prod_id, nom, cant, total, unit in productos_carrito:
                curr = c.execute("SELECT stock FROM productos WHERE id=?", (prod_id,)).fetchone()
                if curr and curr[0] >= cant:
                    nuevo_stock = curr[0] - cant
                    c.execute("UPDATE productos SET stock=? WHERE id=?", (nuevo_stock, prod_id))
                else:
                    raise Exception(f"Stock insuficiente para el producto: {nom}")
                c.execute('''INSERT INTO ventas_productos (cita_id, producto_id, cantidad, precio_unitario, fecha) 
                             VALUES (?,?,?,?,?)''', (id_principal, prod_id, cant, unit, hoy_iso))

            for m, v in pagos:
                if v > 0: 
                    desc_pago = "Venta Servicios"
                    if len(productos_carrito) > 0: desc_pago += " + Productos"
                    c.execute("INSERT INTO pagos (cita_id, metodo, monto, fecha, hora, descripcion_extra) VALUES (?,?,?,?,?,?)", 
                             (id_principal, m, v, hoy_iso, datetime.now().strftime("%H:%M"), desc_pago))
            
            if descuento_dinero > 0:
                c.execute("INSERT INTO gastos (fecha, tipo, categoria, descripcion, metodo, valor) VALUES (?,?,?,?,?,?)", 
                         (hoy_iso, "GASTO", "Descuento Ventas", f"Descuento a {cliente_nombre}", "Cruce Contable", descuento_dinero))
            
            conn.commit()
            return True, "Venta Registrada Correctamente"
        except Exception as e:
            conn.rollback()
            return False, f"Error al procesar cobro: {str(e)}"
        finally:
            conn.close()

    def saldar_cuenta_por_cobrar(self, id_pago, nuevo_metodo):
        conn = self.conectar()
        try:
            c = conn.cursor()
            c.execute("UPDATE pagos SET metodo=?, fecha=?, hora=? WHERE id=?", 
                      (nuevo_metodo, datetime.now().strftime("%Y-%m-%d"), datetime.now().strftime("%H:%M"), id_pago))
            conn.commit()
            return True, "Deuda Saldada"
        except Exception as e:
            conn.rollback()
            return False, str(e)
        finally:
            conn.close()

    def realizar_abono_deuda(self, tipo_deuda, id_registro, monto_abono, metodo_pago):
        conn = self.conectar()
        hoy_iso = datetime.now().strftime("%Y-%m-%d")
        try:
            c = conn.cursor()
            if tipo_deuda == 'CLIENTE':
                deuda = c.execute("SELECT monto, cita_id FROM pagos WHERE id=?", (id_registro,)).fetchone()
                if not deuda: return False, "No existe"
                monto_deuda, cita_id = deuda
                
                if monto_abono >= monto_deuda:
                    c.execute("DELETE FROM pagos WHERE id=?", (id_registro,))
                    c.execute("INSERT INTO pagos (cita_id, metodo, monto, fecha, hora, descripcion_extra) VALUES (?,?,?,?,?,?)",
                                    (cita_id, metodo_pago, monto_deuda, hoy_iso, datetime.now().strftime("%H:%M"), "Pago Deuda Crédito"))
                    conn.commit()
                    return True, "Saldado Total"
                else:
                    nuevo_saldo = monto_deuda - monto_abono
                    c.execute("UPDATE pagos SET monto=? WHERE id=?", (nuevo_saldo, id_registro))
                    c.execute("INSERT INTO pagos (cita_id, metodo, monto, fecha, hora, descripcion_extra) VALUES (?,?,?,?,?,?)",
                                    (cita_id, metodo_pago, monto_abono, hoy_iso, datetime.now().strftime("%H:%M"), "Abono Crédito"))
                    conn.commit()
                    return True, f"Abono OK. Restan: ${nuevo_saldo:,.0f}"

            elif tipo_deuda == 'PROFESIONAL':
                prestamo = c.execute("SELECT monto FROM prestamos WHERE id=?", (id_registro,)).fetchone()
                if not prestamo: return False, "No encontrado"
                monto_prestamo = prestamo[0]
                
                c.execute("INSERT INTO abonos_prestamos (prestamo_id, valor, fecha, descripcion, metodo) VALUES (?,?,?,?,?)",
                                (id_registro, monto_abono, hoy_iso, "Abono Voluntario", metodo_pago))
                
                nuevo_saldo = monto_prestamo - monto_abono
                estado = 'Pagado' if nuevo_saldo <= 0 else 'Pendiente'
                monto_final = 0 if nuevo_saldo <= 0 else nuevo_saldo
                
                c.execute("UPDATE prestamos SET monto=?, estado=? WHERE id=?", (monto_final, estado, id_registro))
                conn.commit()
                return True, "Abono Registrado"

        except Exception as e:
            conn.rollback()
            return False, str(e)
        finally:
            conn.close()

    def pagar_nomina_flexible(self, ids_citas, abono_prestamos, lista_pagos_nomina, profesional_nombre):
        conn = self.conectar()
        hoy_iso = datetime.now().strftime("%Y-%m-%d")
        try:
            cursor = conn.cursor()
            # Buscar ID de empleado en terceros
            pid = cursor.execute("SELECT id FROM terceros WHERE nombre_completo=? AND es_empleado=1", (profesional_nombre,)).fetchone()[0]
            
            for ic in ids_citas: 
                cursor.execute("UPDATE citas SET nomina_pagada = 1 WHERE id = ?", (ic,))
            
            if abono_prestamos > 0:
                prestamos = cursor.execute("SELECT id, monto FROM prestamos WHERE profesional_id=? AND estado='Pendiente' ORDER BY id ASC", (pid,)).fetchall()
                remanente = abono_prestamos
                for pid_p, monto_orig in prestamos:
                    if remanente <= 0: break
                    abono_real = 0
                    if remanente >= monto_orig:
                        cursor.execute("UPDATE prestamos SET estado='Pagado', monto=0 WHERE id=?", (pid_p,))
                        abono_real = monto_orig
                        remanente -= monto_orig
                    else:
                        cursor.execute("UPDATE prestamos SET monto=? WHERE id=?", (monto_orig - remanente, pid_p))
                        abono_real = remanente
                        remanente = 0
                    cursor.execute("INSERT INTO abonos_prestamos (prestamo_id, valor, fecha, descripcion) VALUES (?,?,?,?)", 
                                  (pid_p, abono_real, hoy_iso, "Deducción Nómina"))
            
            for met, val in lista_pagos_nomina:
                if val > 0: 
                    cursor.execute("INSERT INTO gastos (fecha, tipo, categoria, descripcion, metodo, valor) VALUES (?,?,?,?,?,?)", 
                                  (hoy_iso, "GASTO", "Nomina", f"Nomina {profesional_nombre}", met, val))
            
            conn.commit()
            return True, "Nómina Liquidada"
        except Exception as e:
            conn.rollback()
            return False, str(e)
        finally:
            conn.close()

    # ==========================================
    #  CONSULTAS SEGURAS (ADAPTADAS A TERCEROS)
    # ==========================================
    def traer_agenda_filtrada(self, fecha_ui=None, busqueda=None):
        # JOINS actualizados a tabla terceros
        sql = '''SELECT c.id, c.fecha, c.hora_inicio, 
                        t_cli.nombre_completo, t_cli.telefono, 
                        s.nombre, 
                        t_pro.nombre_completo, c.estado 
                 FROM citas c 
                 JOIN terceros t_cli ON c.cliente_id = t_cli.id 
                 JOIN servicios s ON c.servicio_id = s.id 
                 JOIN terceros t_pro ON c.profesional_id = t_pro.id 
                 WHERE 1=1 '''
        params = []
        if fecha_ui: 
            sql += " AND c.fecha = ?"
            params.append(self.f_to_iso(fecha_ui))
        
        if busqueda:
            term = f"%{busqueda}%"
            sql += " AND (t_cli.nombre_completo LIKE ? OR t_cli.telefono LIKE ?)"
            params.extend([term, term])
            
        sql += " AND c.estado != 'Cancelado' ORDER BY c.fecha ASC, c.hora_inicio ASC"
        
        with self.conectar() as conn: 
            rows = conn.execute(sql, params).fetchall()
            return [(r[0], self.f_to_ui(r[1]), r[2], r[3], r[4], r[5], r[6], r[7]) for r in rows]

    def traer_citas_futuras_cliente(self, texto_busqueda):
        busqueda = f"%{texto_busqueda}%"
        sql = '''SELECT c.id, c.fecha, c.hora_inicio, t_cli.nombre_completo, t_cli.telefono, s.nombre, t_pro.nombre_completo, c.estado 
                 FROM citas c 
                 JOIN terceros t_cli ON c.cliente_id = t_cli.id 
                 JOIN servicios s ON c.servicio_id = s.id 
                 JOIN terceros t_pro ON c.profesional_id = t_pro.id 
                 WHERE (t_cli.nombre_completo LIKE ? OR t_cli.telefono LIKE ?) AND c.estado IN ('Pendiente', 'Reagendado')
                 ORDER BY c.fecha ASC, c.hora_inicio ASC'''
        with self.conectar() as conn: 
            rows = conn.execute(sql, (busqueda, busqueda)).fetchall()
            return [(r[0], self.f_to_ui(r[1]), r[2], r[3], r[4], r[5], r[6], r[7]) for r in rows]

    def buscar_deudas_pendientes(self, tipo, texto):
        busqueda = f"%{texto}%"
        with self.conectar() as conn:
            if tipo == 'CLIENTE':
                sql = '''SELECT p.id, p.fecha, t.nombre_completo, t.telefono, p.monto, 'Venta Crédito'
                         FROM pagos p JOIN citas c ON p.cita_id = c.id JOIN terceros t ON c.cliente_id = t.id 
                         WHERE p.metodo = 'CREDITO' AND (t.nombre_completo LIKE ? OR t.telefono LIKE ?) ORDER BY p.fecha DESC'''
                rows = conn.execute(sql, (busqueda, busqueda)).fetchall()
                return [(r[0], self.f_to_ui(r[1]), r[2], r[3], r[4], r[5]) for r in rows]
            elif tipo == 'PROFESIONAL':
                sql = '''SELECT pr.id, pr.fecha, t.nombre_completo, 'N/A', pr.monto, pr.descripcion
                         FROM prestamos pr JOIN terceros t ON pr.profesional_id = t.id
                         WHERE pr.estado = 'Pendiente' AND t.nombre_completo LIKE ? ORDER BY pr.fecha DESC'''
                rows = conn.execute(sql, (busqueda,)).fetchall()
                return [(r[0], self.f_to_ui(r[1]), r[2], r[3], r[4], r[5]) for r in rows]
        return []

    # ... (Se mantienen igual: validar_choque, guardar_paquete_citas pero usando tabla terceros internamente) ...
    def validar_choque(self, fecha_ui, hora_ini, duracion, pro_nombre):
        try:
            fecha_iso = self.f_to_iso(fecha_ui)
            fmt="%H:%M"
            ini = datetime.strptime(hora_ini, fmt)
            fin = ini + timedelta(minutes=int(duracion))
            h_fin = fin.strftime(fmt)
            
            with self.conectar() as conn:
                # Obtener ID desde terceros
                pid = conn.execute("SELECT id FROM terceros WHERE nombre_completo=? AND es_empleado=1", (pro_nombre,)).fetchone()[0]
                
                citas = conn.execute('''SELECT hora_inicio, hora_fin FROM citas 
                       WHERE fecha=? AND profesional_id=? AND estado!='Cancelado' ''', (fecha_iso, pid)).fetchall()
                
                bloqueos = conn.execute('''SELECT hora_inicio, hora_fin, motivo FROM bloqueos 
                       WHERE fecha=? AND (profesional_id=? OR profesional_id=0)''', (fecha_iso, pid)).fetchall()
            
            for ci, cf in citas:
                if (ini < datetime.strptime(cf, fmt)) and (fin > datetime.strptime(ci, fmt)): 
                    return True, f"Ocupado ({ci}-{cf})"
            
            for bi, bf, mot in bloqueos:
                if (ini < datetime.strptime(bf, fmt)) and (fin > datetime.strptime(bi, fmt)): 
                    return True, f"BLOQUEADO: {mot}"
            
            return False, h_fin
        except Exception as e: return True, str(e)

    def guardar_paquete_citas(self, carrito, nom, tel):
        try:
            # Lógica mejorada: Buscar cliente, si no existe, CREAR UNO BÁSICO en Terceros
            cli = self.buscar_cliente(tel)
            if not cli:
                with self.conectar() as conn: 
                    # Insertar en Terceros con flag cliente
                    cur = conn.execute("INSERT INTO terceros (nombre_completo, nombre1, telefono, es_cliente, fecha_registro) VALUES (?,?,?,1,?)", 
                                     (nom, nom, str(tel), datetime.now().strftime("%Y-%m-%d")))
                    cli_id = cur.lastrowid
            else: 
                cli_id = cli[0]
            
            with self.conectar() as conn:
                for i in carrito:
                    pid = conn.execute("SELECT id FROM terceros WHERE nombre_completo=? AND es_empleado=1",(i['profesional'],)).fetchone()[0]
                    sid = conn.execute("SELECT id FROM servicios WHERE nombre=?",(i['servicio'],)).fetchone()[0]
                    f_iso = self.f_to_iso(i['fecha'])
                    
                    conn.execute("INSERT INTO citas (cliente_id, profesional_id, servicio_id, fecha, hora_inicio, hora_fin, precio_final, estado) VALUES (?,?,?,?,?,?,?, 'Pendiente')",
                                 (cli_id, pid, sid, f_iso, i['inicio'], i['fin'], i['precio']))
            return True, "Agendado"
        except Exception as e: return False, str(e)

    def obtener_datos_agenda_visual(self, fecha_ui):
        fecha_iso = self.f_to_iso(fecha_ui)
        data = []
        with self.conectar() as conn:
            citas = conn.execute('''SELECT t_pro.nombre_completo, c.hora_inicio, s.duracion_min, t_cli.nombre_completo, s.nombre
                                   FROM citas c 
                                   JOIN terceros t_pro ON c.profesional_id = t_pro.id
                                   JOIN servicios s ON c.servicio_id = s.id 
                                   JOIN terceros t_cli ON c.cliente_id = t_cli.id
                                   WHERE c.fecha = ? AND c.estado != 'Cancelado' ''', (fecha_iso,)).fetchall()
            for c in citas: data.append(c + ('CITA',))
            
            pros = [x[0] for x in conn.execute("SELECT nombre_completo FROM terceros WHERE es_empleado=1").fetchall()]
            bloqs = conn.execute('''SELECT profesional_id, hora_inicio, hora_fin, motivo FROM bloqueos WHERE fecha=?''', (fecha_iso,)).fetchall()
            
            for pid, hi, hf, mot in bloqs:
                fmt = "%H:%M"
                dur = int((datetime.strptime(hf, fmt) - datetime.strptime(hi, fmt)).total_seconds() / 60)
                if pid == 0: 
                    for p_nombre in pros: data.append((p_nombre, hi, dur, "BLOQUEADO", mot, 'BLOQUEO'))
                else:
                    p_nom = conn.execute("SELECT nombre_completo FROM terceros WHERE id=?", (pid,)).fetchone()[0]
                    data.append((p_nom, hi, dur, "NO DISP.", mot, 'BLOQUEO'))
        return data

    # ==========================================
    #  REPORTES Y EXPORTACIÓN
    # ==========================================
    def obtener_balance_financiero(self, f1_ui, f2_ui):
        d1_iso = self.f_to_iso(f1_ui)
        d2_iso = self.f_to_iso(f2_ui)
        pagos_total = 0; gastos_total = 0; descuentos_total = 0; flujo = {}
        
        with self.conectar() as conn:
            pagos = conn.execute("SELECT metodo, monto FROM pagos WHERE fecha BETWEEN ? AND ?", (d1_iso, d2_iso)).fetchall()
            for met, mon in pagos:
                if met != "CREDITO": 
                    if met not in flujo: flujo[met] = {'in': 0, 'out': 0}
                    flujo[met]['in'] += mon
                pagos_total += mon
            
            gastos = conn.execute("SELECT categoria, metodo, valor FROM gastos WHERE fecha BETWEEN ? AND ?", (d1_iso, d2_iso)).fetchall()
            for cat, met, val in gastos:
                gastos_total += val
                if cat == "Descuento Ventas": descuentos_total += val
                else:
                    if met not in flujo: flujo[met] = {'in': 0, 'out': 0}
                    flujo[met]['out'] += val
        
        venta_bruta = pagos_total + descuentos_total
        utilidad = venta_bruta - gastos_total
        return venta_bruta, gastos_total, utilidad, flujo

    def generar_excel_visual_completo(self, ruta):
        if not HAS_OPENPYXL: return False, "Falta librería openpyxl"
        try:
            wb = Workbook()
            if 'Sheet' in wb.sheetnames: del wb['Sheet']
            
            fill_header = PatternFill("solid", fgColor="D3D3D3")
            fill_cita = PatternFill("solid", fgColor="90EE90")
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            bold_font = Font(bold=True)
            
            with self.conectar() as conn:
                pros = [x[0] for x in conn.execute("SELECT nombre_completo FROM terceros WHERE es_empleado=1 ORDER BY id").fetchall()]
                fechas_citas_iso = [x[0] for x in conn.execute("SELECT DISTINCT fecha FROM citas WHERE estado!='Cancelado'").fetchall()]
                fechas_unicas_iso = sorted(list(set(fechas_citas_iso)))
                
                if not fechas_unicas_iso:
                    ws = wb.create_sheet("Sin Citas")
                    ws['A1'] = "No hay citas registradas"
                    wb.save(ruta)
                    return True, f"Excel vacío guardado en {ruta}"
                
                for fecha_iso in fechas_unicas_iso:
                    fecha_ui = self.f_to_ui(fecha_iso)
                    ws = wb.create_sheet(fecha_ui)
                    
                    ws.cell(row=1, column=1, value="HORA").font = bold_font
                    ws.cell(row=1, column=1).fill = fill_header
                    ws.column_dimensions['A'].width = 10
                    
                    col_map = {}
                    for i, p in enumerate(pros):
                        col_idx = i + 2
                        c = ws.cell(row=1, column=col_idx, value=p.upper())
                        c.font = bold_font; c.fill = fill_header; c.alignment = center_align; c.border = thin_border
                        ws.column_dimensions[chr(64 + col_idx)].width = 25
                        col_map[p] = col_idx
                    
                    time_map = {}
                    row_idx = 2
                    t_start = datetime.strptime("06:00", "%H:%M")
                    t_end = datetime.strptime("22:00", "%H:%M")
                    curr_t = t_start
                    while curr_t <= t_end:
                        hora_str = curr_t.strftime("%H:%M")
                        cell = ws.cell(row=row_idx, column=1, value=hora_str)
                        cell.border = thin_border
                        time_map[hora_str] = row_idx
                        curr_t += timedelta(minutes=20)
                        row_idx += 1
                    
                    citas = conn.execute('''SELECT t_pro.nombre_completo, c.hora_inicio, s.duracion_min, t_cli.nombre_completo, s.nombre, c.precio_final
                                            FROM citas c 
                                            JOIN terceros t_pro ON c.profesional_id = t_pro.id 
                                            JOIN servicios s ON c.servicio_id = s.id 
                                            JOIN terceros t_cli ON c.cliente_id = t_cli.id
                                            WHERE c.fecha = ? AND c.estado != 'Cancelado' ''', (fecha_iso,)).fetchall()
                    
                    for pro_nom, h_ini, dur, cli_nom, srv_nom, precio in citas:
                        if pro_nom in col_map:
                            bloques = int(dur / 20)
                            if bloques < 1: bloques = 1
                            if h_ini in time_map:
                                r_start = time_map[h_ini]
                                c_idx = col_map[pro_nom]
                                for b in range(bloques):
                                    if (r_start + b) < row_idx:
                                        cell = ws.cell(row=r_start + b, column=c_idx)
                                        cell.fill = fill_cita; cell.border = thin_border
                                        if b == 0: 
                                            cell.value = f"{cli_nom}\n{srv_nom}"
                                            cell.alignment = center_align
            
            wb.save(ruta)
            return True, f"Libro generado: {ruta}"
        except Exception as e: return False, str(e)

    # ==========================================
    #  CRUD Y FUNCIONES MENORES
    # ==========================================
    def crear_servicio(self, n, d, p):
        return self.crear_item_unificado('SERVICIO', n, p, d)

    def eliminar_servicio(self, n):
        try:
            with self.conectar() as conn: conn.execute("DELETE FROM servicios WHERE nombre=?",(n,))
            return True, "Eliminado"
        except: return False, "Error"

    def crear_profesional(self, n, c, s):
        # Mantenido para compatibilidad: Crea un Tercero marcado como empleado
        # nombre 'n' se asume como nombre completo
        datos = {
            'n1': n, 'es_empleado': True, 'comision': c, 'servicios': s,
            'fecha_registro': datetime.now().strftime("%Y-%m-%d")
        }
        return self.guardar_tercero(datos)

    def eliminar_profesional(self, n):
        # Eliminar el registro de tercero con ese nombre
        try:
            with self.conectar() as conn: conn.execute("DELETE FROM terceros WHERE nombre_completo=? AND es_empleado=1",(n,))
            return True, "Eliminado"
        except: return False, "Error"

    def crear_medio_pago(self, nombre):
        try:
            with self.conectar() as conn: conn.execute("INSERT INTO medios_pago (nombre) VALUES (?)", (nombre,))
            return True, "Creado"
        except Exception as e: return False, str(e)

    def eliminar_medio_pago(self, nombre):
        if nombre == "CREDITO": return False, "No se puede borrar CREDITO"
        try:
            with self.conectar() as conn: conn.execute("DELETE FROM medios_pago WHERE nombre=?", (nombre,))
            return True, "Eliminado"
        except: return False, "Error"

    def editar_cliente(self, id_cli, nombre, tel):
        # Adaptado a terceros
        try:
            with self.conectar() as conn: conn.execute("UPDATE terceros SET nombre_completo=?, telefono=? WHERE id=?", (nombre, tel, id_cli))
            return True, "Actualizado"
        except Exception as e: return False, str(e)

    def importar_clientes_masivo(self, ruta_archivo):
        try:
            df = pd.read_excel(ruta_archivo)
            if 'Nombre' not in df.columns or 'Telefono' not in df.columns: return False, "Requiere columnas Nombre y Telefono"
            count = 0
            with self.conectar() as conn:
                for _, row in df.iterrows():
                    try:
                        conn.execute("INSERT INTO terceros (nombre_completo, nombre1, telefono, es_cliente, fecha_registro) VALUES (?,?,?,1,?)", 
                                     (str(row['Nombre']), str(row['Nombre']), str(row['Telefono']), datetime.now().strftime("%Y-%m-%d")))
                        count += 1
                    except sqlite3.IntegrityError: pass
            return True, f"Importados {count}"
        except Exception as e: return False, str(e)

    def reagendar_cita(self, id_cita, nueva_fecha_ui, nueva_hora, nuevo_pro):
        with self.conectar() as conn:
            cita = conn.execute('''SELECT s.duracion_min FROM citas c JOIN servicios s ON c.servicio_id=s.id WHERE c.id=?''', (id_cita,)).fetchone()
            if not cita: return False, "Cita no existe"
            dur = cita[0]
        ocupado, fin_o_msg = self.validar_choque(nueva_fecha_ui, nueva_hora, dur, nuevo_pro)
        if ocupado: return False, fin_o_msg
        
        nueva_fecha_iso = self.f_to_iso(nueva_fecha_ui)
        try:
            with self.conectar() as conn:
                pid = conn.execute("SELECT id FROM terceros WHERE nombre_completo=? AND es_empleado=1", (nuevo_pro,)).fetchone()[0]
                conn.execute("UPDATE citas SET fecha=?, hora_inicio=?, hora_fin=?, profesional_id=?, estado='Reagendado' WHERE id=?", 
                             (nueva_fecha_iso, nueva_hora, fin_o_msg, pid, id_cita))
            return True, "Cita Reagendada"
        except Exception as e: return False, str(e)

    def cancelar_cita(self, id_cita):
        try:
            with self.conectar() as conn: conn.execute("UPDATE citas SET estado = 'Cancelado' WHERE id = ?", (id_cita,))
            return True, "Cita Cancelada"
        except Exception as e: return False, str(e)

    def crear_bloqueo(self, pro_nombre, f_inicio, f_fin, h_ini, h_fin, motivo):
        try:
            d_ini = datetime.strptime(f_inicio, "%d-%m-%y"); d_fin = datetime.strptime(f_fin, "%d-%m-%y")
            with self.conectar() as conn:
                if pro_nombre == "TODOS": pid = 0
                else: pid = conn.execute("SELECT id FROM terceros WHERE nombre_completo=? AND es_empleado=1", (pro_nombre,)).fetchone()[0]
                delta = d_fin - d_ini
                for i in range(delta.days + 1):
                    dia_actual = d_ini + timedelta(days=i); fecha_iso = dia_actual.strftime("%Y-%m-%d")
                    conn.execute("INSERT INTO bloqueos (profesional_id, fecha, hora_inicio, hora_fin, motivo) VALUES (?,?,?,?,?)",
                                 (pid, fecha_iso, h_ini, h_fin, motivo))
            return True, "Bloqueo Creado"
        except Exception as e: return False, str(e)

    def establecer_horario_global(self, fecha_ui, h_apertura, h_cierre):
        fecha_iso = self.f_to_iso(fecha_ui)
        try:
            with self.conectar() as conn:
                conn.execute("DELETE FROM bloqueos WHERE fecha=? AND profesional_id=0 AND motivo='Fuera de Horario'", (fecha_iso,))
                if h_apertura != "00:00": conn.execute("INSERT INTO bloqueos (profesional_id, fecha, hora_inicio, hora_fin, motivo) VALUES (0, ?, '00:00', ?, 'Fuera de Horario')", (fecha_iso, h_apertura))
                if h_cierre != "23:59": conn.execute("INSERT INTO bloqueos (profesional_id, fecha, hora_inicio, hora_fin, motivo) VALUES (0, ?, ?, '23:59', 'Fuera de Horario')", (fecha_iso, h_cierre))
            return True, f"Horario establecido para {fecha_ui}"
        except Exception as e: return False, str(e)

    def eliminar_bloqueo(self, id_bloqueo):
        try:
            with self.conectar() as conn: conn.execute("DELETE FROM bloqueos WHERE id=?", (id_bloqueo,))
            return True, "Desbloqueado"
        except Exception as e: return False, str(e)

    def crear_prestamo(self, profesional, monto, desc):
        hoy_iso = datetime.now().strftime("%Y-%m-%d")
        try:
            with self.conectar() as conn:
                pid = conn.execute("SELECT id FROM terceros WHERE nombre_completo=? AND es_empleado=1", (profesional,)).fetchone()[0]
                conn.execute("INSERT INTO prestamos (profesional_id, monto, fecha, descripcion) VALUES (?,?,?,?)", (pid, float(monto), hoy_iso, desc))
                conn.execute("INSERT INTO gastos (fecha, tipo, categoria, descripcion, metodo, valor) VALUES (?,?,?,?,?,?)",
                             (hoy_iso, "GASTO", "Prestamos", f"Prestamo a {profesional}", "Efectivo", float(monto)))
            return True, "Préstamo registrado"
        except Exception as e: return False, str(e)
    
    # Metodos de consulta simples
    def get_info_servicio(self, n):
        with self.conectar() as conn: return conn.execute("SELECT id, duracion_min, precio FROM servicios WHERE nombre=?",(n,)).fetchone()
    def traer_dias_ocupados(self):
        with self.conectar() as conn: 
            fechas_iso = [x[0] for x in conn.execute("SELECT DISTINCT fecha FROM citas WHERE estado!='Cancelado'").fetchall()]
            return [self.f_to_ui(f) for f in fechas_iso]
    def traer_intervalos_ocupados(self, fecha_ui, pro_nombre):
        fecha_iso = self.f_to_iso(fecha_ui); intervalos = []
        with self.conectar() as conn:
            pid = conn.execute("SELECT id FROM terceros WHERE nombre_completo=? AND es_empleado=1", (pro_nombre,)).fetchone()[0]
            citas = conn.execute('''SELECT hora_inicio, hora_fin FROM citas WHERE fecha = ? AND profesional_id = ? AND estado != 'Cancelado' ''', (fecha_iso, pid)).fetchall()
            intervalos.extend(citas)
            bloqs = conn.execute('''SELECT hora_inicio, hora_fin FROM bloqueos WHERE fecha=? AND (profesional_id=? OR profesional_id=0)''', (fecha_iso, pid)).fetchall()
            intervalos.extend(bloqs)
        return intervalos
    def confirm_asistencia(self, id_c):
        with self.conectar() as conn: conn.execute("UPDATE citas SET estado='Por Cobrar' WHERE id=?",(id_c,))
    def confirmar_asistencia(self, id_c):
        with self.conectar() as conn: conn.execute("UPDATE citas SET estado='Por Cobrar' WHERE id=?",(id_c,))
    def traer_por_cobrar(self):
        sql = '''SELECT t.id, t.nombre_completo, t.telefono, GROUP_CONCAT(c.id), SUM(c.precio_final), GROUP_CONCAT(s.nombre)
                 FROM citas c JOIN terceros t ON c.cliente_id=t.id JOIN servicios s ON c.servicio_id=s.id
                 WHERE c.estado='Por Cobrar' GROUP BY t.id'''
        with self.conectar() as conn: return conn.execute(sql).fetchall()
    def traer_lista_bloqueos(self):
        with self.conectar() as conn:
            rows = conn.execute('''SELECT b.id, CASE WHEN b.profesional_id = 0 THEN 'TODOS' ELSE t.nombre_completo END, b.fecha, b.hora_inicio, b.hora_fin, b.motivo
                                   FROM bloqueos b LEFT JOIN terceros t ON b.profesional_id = t.id 
                                   ORDER BY b.fecha DESC LIMIT 50''').fetchall()
            return [(r[0], r[1], self.f_to_ui(r[2]), r[3], r[4], r[5]) for r in rows]
    def traer_profesionales_habilitados_por_fecha(self, fecha_ui):
        fecha_iso = self.f_to_iso(fecha_ui)
        with self.conectar() as conn:
            todos = conn.execute("SELECT id, nombre_completo FROM terceros WHERE es_empleado=1").fetchall()
            bloqueos = conn.execute("SELECT profesional_id, hora_inicio, hora_fin FROM bloqueos WHERE fecha=?", (fecha_iso,)).fetchall()
            for bid, ini, fin in bloqueos:
                if bid == 0 and ((ini=="00:00" and fin=="23:59") or (ini=="00:00" and fin=="00:00")): return []
            disponibles = []
            for pid, nom in todos:
                bloqueado = False
                for bid, ini, fin in bloqueos:
                    if bid == pid and ((ini=="00:00" and fin=="23:59") or (ini=="00:00" and fin=="00:00")):
                        bloqueado = True; break
                if not bloqueado: disponibles.append(nom)
            return disponibles
    def traer_info_liquidacion(self, profesional, f1_str, f2_str):
        d1_iso = self.f_to_iso(f1_str); d2_iso = self.f_to_iso(f2_str); ventas=[]; prestamos=[]
        with self.conectar() as conn:
            raw_v = conn.execute('''SELECT c.id, c.fecha, t_cli.nombre_completo, s.nombre, c.precio_final, t_pro.comision 
                                    FROM citas c 
                                    JOIN terceros t_pro ON c.profesional_id = t_pro.id 
                                    JOIN terceros t_cli ON c.cliente_id = t_cli.id 
                                    JOIN servicios s ON c.servicio_id = s.id
                                    WHERE t_pro.nombre_completo = ? AND c.estado = 'Pagado' AND c.nomina_pagada = 0 AND c.fecha BETWEEN ? AND ?''', 
                                    (profesional, d1_iso, d2_iso)).fetchall()
            for r in raw_v: ventas.append((r[0], self.f_to_ui(r[1]), r[2], r[3], r[4], r[4]*(r[5]/100))) 
            raw_p = conn.execute('''SELECT pr.id, pr.fecha, pr.monto, pr.descripcion FROM prestamos pr JOIN terceros t ON pr.profesional_id = t.id
                                        WHERE t.nombre_completo = ? AND pr.estado = 'Pendiente' ''', (profesional,)).fetchall()
            for r in raw_p: prestamos.append((r[0], self.f_to_ui(r[1]), r[2], r[3]))
        return ventas, prestamos
    def traer_estado_cuenta_prestamos(self, profesional):
        with self.conectar() as conn:
            pid = conn.execute("SELECT id FROM terceros WHERE nombre_completo=? AND es_empleado=1", (profesional,)).fetchone()[0]
            rows = conn.execute("SELECT fecha, descripcion, monto, estado FROM prestamos WHERE profesional_id = ? ORDER BY id DESC", (pid,)).fetchall()
            return [(self.f_to_ui(r[0]), r[1], r[2], r[3]) for r in rows]
    def traer_kardex_prestamos(self, profesional):
        try:
            with self.conectar() as conn:
                pid = conn.execute("SELECT id FROM terceros WHERE nombre_completo=? AND es_empleado=1", (profesional,)).fetchone()
                if not pid: return []
                pid = pid[0]
                prestamos = conn.execute("SELECT fecha, descripcion, monto, 'PRESTAMO (+)', id FROM prestamos WHERE profesional_id = ?", (pid,)).fetchall()
                abonos = conn.execute("SELECT a.fecha, a.descripcion, a.valor, 'ABONO (-)', p.id FROM abonos_prestamos a JOIN prestamos p ON a.prestamo_id = p.id WHERE p.profesional_id = ?", (pid,)).fetchall()
                movs = []
                for p in prestamos: movs.append({'f': p[0], 'd': p[1], 'm': p[2], 't': 'PRESTAMO (+)', 'dt': datetime.strptime(p[0], "%Y-%m-%d")})
                for a in abonos: movs.append({'f': a[0], 'd': a[1], 'm': a[2], 't': 'ABONO (-)', 'dt': datetime.strptime(a[0], "%Y-%m-%d")})
                movs.sort(key=lambda x: x['dt']); kardex = []; saldo = 0
                for m in movs:
                    saldo += m['m'] if m['t'] == 'PRESTAMO (+)' else -m['m']
                    kardex.append((self.f_to_ui(m['f']), m['t'], m['d'], f"${m['m']:,.0f}", f"${saldo:,.0f}"))
                return kardex
        except: return []
    def traer_historial_ventas_por_fecha(self, f1, f2):
        d1_iso = self.f_to_iso(f1); d2_iso = self.f_to_iso(f2)
        with self.conectar() as conn:
            rows = conn.execute('''SELECT p.fecha, t.nombre_completo, s.nombre, p.monto, p.metodo 
                                   FROM pagos p JOIN citas c ON p.cita_id = c.id 
                                   JOIN terceros t ON c.cliente_id = t.id 
                                   JOIN servicios s ON c.servicio_id = s.id 
                                   WHERE p.fecha BETWEEN ? AND ? ORDER BY p.fecha DESC''', (d1_iso, d2_iso)).fetchall()
            return [(self.f_to_ui(r[0]), r[1], r[2], r[3], r[4]) for r in rows]
    def traer_historial_gastos_por_fecha(self, f1, f2):
        d1_iso = self.f_to_iso(f1); d2_iso = self.f_to_iso(f2)
        with self.conectar() as conn:
            rows = conn.execute("SELECT fecha, tipo, categoria, descripcion, metodo, valor FROM gastos WHERE fecha BETWEEN ? AND ? ORDER BY fecha DESC", (d1_iso, d2_iso)).fetchall()
            return [(self.f_to_ui(r[0]), r[1], r[2], r[3], r[4], r[5]) for r in rows]
    def traer_detalle_ventas(self, f1, f2): return self.traer_historial_ventas_por_fecha(f1, f2)
    def traer_detalle_gastos(self, f1, f2): return self.traer_historial_gastos_por_fecha(f1, f2)
    def registrar_gasto(self, t, c, d, m, v):
        with self.conectar() as conn: conn.execute("INSERT INTO gastos (fecha, tipo, categoria, descripcion, metodo, valor) VALUES (?,?,?,?,?,?)", (datetime.now().strftime("%Y-%m-%d"), t, c, d, m, v))
        return True
    def traer_medios_pago(self):
        with self.conectar() as conn: return [x[0] for x in conn.execute("SELECT nombre FROM medios_pago").fetchall()]
    def obtener_cierre_caja_dia(self, fecha_ui):
        fecha_iso = self.f_to_iso(fecha_ui)
        res = {'ventas': {}, 'gastos': {}, 'prestamos': {}, 'detalle_gastos_lista': [], 'detalle_creditos_lista': [], 'abonos_prestamos': {}}
        with self.conectar() as conn:
            pagos = conn.execute('''SELECT p.metodo, p.monto, t.nombre_completo FROM pagos p JOIN citas c ON p.cita_id = c.id JOIN terceros t ON c.cliente_id = t.id WHERE p.fecha=?''', (fecha_iso,)).fetchall()
            for met, mon, cli_nom in pagos:
                res['ventas'][met] = res['ventas'].get(met, 0) + mon
                if met == 'CREDITO': res['detalle_creditos_lista'].append((f"Crédito: {cli_nom}", mon))
            try:
                abonos = conn.execute("SELECT metodo, valor FROM abonos_prestamos WHERE fecha=?", (fecha_iso,)).fetchall()
                for met, val in abonos:
                    m = met if met else "Efectivo"
                    res['abonos_prestamos'][m] = res['abonos_prestamos'].get(m, 0) + val
            except: pass
            gastos = conn.execute("SELECT categoria, metodo, valor, descripcion FROM gastos WHERE fecha=?", (fecha_iso,)).fetchall()
            for cat, met, val, desc in gastos:
                if cat == 'Prestamos': res['prestamos'][met] = res['prestamos'].get(met, 0) + val
                elif cat == 'Descuento Ventas': pass
                else:
                    res['gastos'][met] = res['gastos'].get(met, 0) + val
                    res['detalle_gastos_lista'].append((desc, met, val))
        return res
    def traer_gastos(self): return self.traer_historial_gastos_por_fecha(datetime.now().strftime("%d-%m-%y"), datetime.now().strftime("%d-%m-%y"))
    def guardar_mensaje_wa(self, mensaje):
        try:
            with self.conectar() as conn: conn.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('msg_wa', ?)", (mensaje,))
            self.registrar_auditoria("CONFIG_WA", "Se actualizó plantilla mensaje WA")
            return True, "Mensaje actualizado"
        except Exception as e: return False, str(e)
    def traer_mensaje_wa(self):
        with self.conectar() as conn:
            res = conn.execute("SELECT valor FROM configuracion WHERE clave='msg_wa'").fetchone()
            return res[0] if res else ""
    def obtener_cita_full(self, id_cita):
        with self.conectar() as conn:
            row = conn.execute('''SELECT c.fecha, c.hora_inicio, t_cli.nombre_completo, t_cli.telefono, s.nombre, t_pro.nombre_completo, c.precio_final
                     FROM citas c 
                     JOIN terceros t_cli ON c.cliente_id = t_cli.id 
                     JOIN servicios s ON c.servicio_id = s.id 
                     JOIN terceros t_pro ON c.profesional_id = t_pro.id 
                     WHERE c.id = ?''', (id_cita,)).fetchone()
            if row: return (self.f_to_ui(row[0]), row[1], row[2], row[3], row[4], row[5], row[6])
            return None
    def exportar_lista_a_excel(self, datos, columnas, ruta):
        if not HAS_OPENPYXL: return False, "Falta openpyxl"
        try:
            df = pd.DataFrame(datos, columns=columnas); df.to_excel(ruta, index=False)
            return True, f"Guardado en {ruta}"
        except Exception as e: return False, str(e)
    def traer_cuentas_por_pagar_comisiones(self):
        with self.conectar() as conn:
            return conn.execute('''SELECT t.nombre_completo, SUM(c.precio_final * (t.comision / 100)) as deuda
                FROM citas c JOIN terceros t ON c.profesional_id = t.id
                WHERE c.estado = 'Pagado' AND c.nomina_pagada = 0 GROUP BY t.nombre_completo''').fetchall()

    def traer_cuentas_por_cobrar_creditos(self):
        sql = '''SELECT p.id, p.fecha, t.nombre_completo, t.telefono, p.monto 
                 FROM pagos p 
                 JOIN citas c ON p.cita_id = c.id 
                 JOIN terceros t ON c.cliente_id = t.id 
                 WHERE p.metodo = 'CREDITO' AND p.monto > 0 
                 ORDER BY p.fecha ASC'''
        with self.conectar() as conn:
            rows = conn.execute(sql).fetchall()
            return [(r[0], self.f_to_ui(r[1]), r[2], r[3], r[4]) for r in rows]

    # ==========================================
    #  MODIFICADO: TICKET CON DATOS REALES
    # ==========================================
    def generar_ticket_termico(self, datos_cita):
        if not HAS_REPORTLAB: return False, "Falta librería reportlab"
        
        # Traer datos reales de la empresa
        emp = self.traer_datos_empresa()
        
        ANCHO_PAPEL = 80 * mm
        # Calculamos alto dinámico aprox según items
        alto_extra = len(datos_cita['items']) * 10 * mm
        ALTO_PAPEL = (180 * mm) + alto_extra
        
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        nombre_pdf = f"Ticket_{timestamp}.pdf"
        
        try:
            c = canvas.Canvas(nombre_pdf, pagesize=(ANCHO_PAPEL, ALTO_PAPEL))
            y = ALTO_PAPEL - 5 * mm
            center_x = ANCHO_PAPEL / 2
            
            # 1. LOGO (Si existe logo.png)
            if os.path.exists("logo.png"):
                try:
                    # Dibujar logo centrado
                    c.drawImage("logo.png", center_x - 12*mm, y - 25*mm, width=24*mm, height=24*mm, preserveAspectRatio=True, mask='auto')
                    y -= 30 * mm
                except: pass
            else:
                y -= 5 * mm

            # 2. ENCABEZADO EMPRESA (Datos dinámicos)
            c.setFont("Helvetica-Bold", 11)
            c.drawCentredString(center_x, y, emp['nombre'])
            y -= 5 * mm
            
            c.setFont("Helvetica", 8)
            if emp['nit']: 
                c.drawCentredString(center_x, y, f"NIT/RUT: {emp['nit']}")
                y -= 4 * mm
            if emp['dir']: 
                c.drawCentredString(center_x, y, emp['dir'])
                y -= 4 * mm
            if emp['tel']: 
                c.drawCentredString(center_x, y, f"Tel: {emp['tel']}")
                y -= 6 * mm
            
            c.line(2*mm, y, ANCHO_PAPEL - 2*mm, y)
            y -= 5 * mm

            # 3. DATOS CITA
            c.setFont("Helvetica-Bold", 9)
            c.drawCentredString(center_x, y, "COMPROBANTE DE VENTA")
            y -= 5 * mm
            c.setFont("Helvetica", 8)
            c.drawString(5*mm, y, f"Fecha: {datos_cita['fecha']}")
            y -= 4 * mm
            c.drawString(5*mm, y, f"Cliente: {datos_cita['cliente']}")
            y -= 6 * mm

            # 4. ITEMS
            c.line(2*mm, y, ANCHO_PAPEL - 2*mm, y)
            y -= 4 * mm
            c.setFont("Helvetica-Bold", 8)
            c.drawString(2*mm, y, "Cant")
            c.drawString(12*mm, y, "Descripción")
            c.drawRightString(ANCHO_PAPEL - 5*mm, y, "Total")
            y -= 4 * mm
            
            c.setFont("Helvetica", 8)
            for item in datos_cita['items']:
                # item viene como (servicio, profesional, hora, precio_str)
                # Ojo: Adaptar si envias items de productos
                desc = item[0]
                precio = item[3]
                
                # Cortar texto largo
                if len(desc) > 22: desc = desc[:22] + "..."
                
                c.drawString(2*mm, y, "1")
                c.drawString(12*mm, y, desc)
                c.drawRightString(ANCHO_PAPEL - 5*mm, y, precio)
                y -= 4 * mm
            
            y -= 2 * mm
            c.line(2*mm, y, ANCHO_PAPEL - 2*mm, y)
            y -= 6 * mm
            
            # 5. TOTAL
            c.setFont("Helvetica-Bold", 14)
            c.drawRightString(ANCHO_PAPEL - 5*mm, y, f"TOTAL: ${datos_cita['total']:,.0f}")
            y -= 8 * mm
            
            c.setFont("Helvetica", 7)
            c.drawCentredString(center_x, y, "¡Gracias por su compra!")
            c.drawCentredString(center_x, y - 3*mm, "Software desarrollado por Tu Empresa")
            
            c.save()
            
            try: os.startfile(nombre_pdf)
            except: pass
            
            return True, "Ticket Generado"
        except Exception as e:
            return False, str(e)